# необходимо установить pip install pandas, pip install openpyxl, pip install gspread, pip install gspread_dataframe==3.2.2
# эксель таблицы загружены в бд sqlite, данная база данных также находится в репозитории проекта

# import os
# from pandas import DataFrame, to_datetime
# import datetime as dt
import sqlite3
import openpyxl
import pandas as pd

'''Экспорт данных из xlsx в sqlite'''
def export_to_sqlite(xl='sber_data.xlsx', lst_xl='', column=''):

    # 1. Создание и подключение к базе

    # Имя базы
    base_name = 'sber_data.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect(base_name)
    # курсор - это специальный объект, который делает запросы и получает результаты запросов
    cursor = connect.cursor()

    # создание таблицы если ее не существует
    cursor.execute(f'CREATE TABLE IF NOT EXISTS {lst_xl} {column}')

    # 2. Работа c xlsx файлом

    # Читаем файл и лист1 книги excel
    file_to_read = openpyxl.load_workbook(xl, data_only=True)
    sheet = file_to_read[lst_xl]

    # Цикл по строкам начиная со второй (в первой заголовки)
    for row in range(2, sheet.max_row + 1):
        # Объявление списка
        data = []
        # Цикл по столбцам начиная 1
        for col in range(1, len(column) + 1):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
        print(data)

        # 3. Запись в базу и закрытие соединения

        # Вставка данных в поля таблицы, необходимо менять количество записей вручную соответственно количеству столбцов
        cursor.execute(f"INSERT INTO {lst_xl} VALUES (?, ?, ?);",
                       (data[0], data[1], data[2].date()))

    # data[1].date()

    # сохраняем изменения
    connect.commit()
    # закрытие соединения
    connect.close()


'''Очистка базы sqlite'''
def clear_base(table):

    # Имя базы
    base_name = 's.sqlite3'

    connect = sqlite3.connect(base_name)
    cursor = connect.cursor()

    # Запись в базу, сохранение и закрытие соединения
    cursor.execute(f"DELETE FROM {table}")
    connect.commit()
    connect.close()


'''SQL запрос возвращает таблицу'''
def input_base(sql_request, base_name='sber_data.sqlite3'):

    # соединение с базой
    connect = sqlite3.connect(base_name)

    # Запрос в базу, сохранение и закрытие соединения
    df = pd.read_sql(sql_request, connect)
    connect.commit()
    connect.close()
    return df


'''переменные для экспорта таблиц эксель в бд'''
# переменные с наименованиями вкладок ексель соответственные названию страниц
transactions, rates = 'Transactions', 'Rates'
vsp_oper_data = 'VSP_oper_data'
distance_metric = 'distance_metric'
users_position = 'users_position'

# переменные с наименованиями столбцов в виде картежей по второму заданию, таблицы по первому заданию заводил немного иначе
transactions_column = ('Client_id', 'Report_date', 'Txn_amount')
rates_column = ('Report_date', 'Ccy_code', 'CCy_rate')
vsp_oper_data_column = ('Client_id', 'Report_date', 'VSP_Number', 'Txn_type', 'Txn_amount')
distance_metric_column = ('VSP', 'VAL', 'VSP_E', 'GROUP_VSP')
users_position_column = ('user_id', 'user_position', 'date_position')


# Запуск функции
if __name__ == '__main__':
    pass
    # input_base(sql_request=zd_1_3)
    # export_to_sqlite(lst_xl=users_position, column=users_position_column)
    # clear_base(table=p_1)
